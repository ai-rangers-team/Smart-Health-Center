"""Bulk-onboard districts + their first district_admin from a CSV or Excel file.

Required columns: district_id, name, state, admin_email — one row per district.
Run: python -m scripts.bulk_onboard_districts scripts/bulk_onboarding.xlsx   (or .xlsx)
  (from api/, with .env + service account configured)
"""
import csv
import sys

from app.seed.demo_data import provision_account
from app.services.language import default_language_for_state

REQUIRED_COLUMNS = ["district_id", "name", "state", "admin_email"]


def _db():
    # Lazy import: keeps this module (and its parsing helpers) importable
    # without live credentials — same convention as app/routers/*.py.
    from app.firestore_client import db
    return db


def _read_csv(path):
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def _read_xlsx(path):
    from openpyxl import load_workbook
    ws = load_workbook(path, read_only=True).active
    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() for h in next(rows)]
    return [dict(zip(header, row)) for row in rows if any(c is not None for c in row)]


def onboard_district(district_id: str, name: str, state: str, admin_email: str):
    dref = _db().collection("districts").document(district_id)
    created = not dref.get().exists
    if created:
        dref.set({
            "name": name, "state": state, "total_centres": 0,
            "default_language": default_language_for_state(state),
        })
    provisioned = provision_account(admin_email, "district_admin", district_id, None)
    return created, provisioned


def main(path: str):
    rows = _read_xlsx(path) if path.lower().endswith(".xlsx") else _read_csv(path)
    if rows and (missing := [c for c in REQUIRED_COLUMNS if c not in rows[0]]):
        print(f"missing required column(s): {', '.join(missing)}")
        sys.exit(1)

    for i, row in enumerate(rows, start=2):  # row 1 is the header
        values = {k: str(row.get(k) or "").strip() for k in REQUIRED_COLUMNS}
        if not all(values.values()):
            print(f"row {i}: skipped — missing a required value")
            continue
        created, provisioned = onboard_district(
            values["district_id"], values["name"], values["state"], values["admin_email"])
        print(f"row {i}: {values['district_id']} — "
              f"district {'created' if created else 'already existed'}, "
              f"admin {'granted' if provisioned else 'pending (never signed in)'}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("usage: python -m scripts.bulk_onboard_districts <path/to/file.csv|.xlsx>")
        sys.exit(1)
    main(sys.argv[1])
